"""
Import/export for reusable payroll profiles.

Supported formats:
  - JSON bundle file
  - CSV directory with one file per profile type
  - Excel workbook with one sheet per profile type
"""
from __future__ import annotations

import csv
import json
from dataclasses import fields
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from models.profile_store import (
    PROFILE_ROOT,
    CompanyProfile,
    DeductionDefaultsProfile,
    EmployeeProfile,
    PayrollAssignmentProfile,
    TaxDefaultsProfile,
    list_profiles,
    load_profiles_by_type,
    profile_to_dict,
    save_company_profile,
    save_deduction_defaults_profile,
    save_employee_profile,
    save_payroll_assignment_profile,
    save_tax_defaults_profile,
)
from models.pay_period import PayFrequency
from models.payroll_calculator import BenefitLine, DeductionLine, EarningLine, FilingStatus


PROFILE_TYPE_KEYS = {
    "company": "companies",
    "employee": "employees",
    "tax": "tax_defaults",
    "deduction": "deduction_defaults",
    "assignment": "assignments",
}

PROFILE_CLASSES = {
    "company": CompanyProfile,
    "employee": EmployeeProfile,
    "tax": TaxDefaultsProfile,
    "deduction": DeductionDefaultsProfile,
    "assignment": PayrollAssignmentProfile,
}

PROFILE_SAVERS = {
    "company": save_company_profile,
    "employee": save_employee_profile,
    "tax": save_tax_defaults_profile,
    "deduction": save_deduction_defaults_profile,
    "assignment": save_payroll_assignment_profile,
}

NESTED_LIST_FACTORIES = {
    "employee": {
        "earnings": EarningLine,
        "other_benefits": BenefitLine,
        "important_notes": None,
    },
    "deduction": {
        "pre_tax_deductions": DeductionLine,
        "post_tax_deductions": DeductionLine,
    },
}

ENUM_FIELDS = {
    "tax": {
        "filing_status": FilingStatus,
        "frequency": PayFrequency,
    },
}

FLOAT_FIELDS = {
    "tax": {"additional_federal_wh", "state_tax_rate_override", "local_tax_rate"},
    "assignment": set(),
}

INT_FIELDS = {
    "tax": {"allowances"},
    "assignment": {"payroll_check_number_start"},
}


def _bundle_from_root(root: str | Path = PROFILE_ROOT) -> dict[str, list[dict[str, Any]]]:
    bundle: dict[str, list[dict[str, Any]]] = {}
    for profile_type, key in PROFILE_TYPE_KEYS.items():
        bundle[key] = [profile_to_dict(profile) for profile in load_profiles_by_type(profile_type, root=root)]
    return bundle


def _normalize_scalar(value: Any) -> Any:
    if value in ("", None):
        return None
    return value


def _serialize_record_for_tabular(record: dict[str, Any]) -> dict[str, Any]:
    serialized: dict[str, Any] = {}
    for key, value in record.items():
        if isinstance(value, (list, dict)):
            serialized[key] = json.dumps(value)
        elif value is None:
            serialized[key] = ""
        else:
            serialized[key] = value
    return serialized


def _deserialize_nested(profile_type: str, record: dict[str, Any]) -> dict[str, Any]:
    nested_fields = NESTED_LIST_FACTORIES.get(profile_type, {})
    enum_fields = ENUM_FIELDS.get(profile_type, {})
    converted = dict(record)

    for key, enum_class in enum_fields.items():
        value = converted.get(key)
        if value in ("", None):
            continue
        converted[key] = enum_class(value)

    for key, factory in nested_fields.items():
        raw_value = converted.get(key, "")
        if raw_value in ("", None):
            converted[key] = []
            continue
        parsed = json.loads(raw_value) if isinstance(raw_value, str) else raw_value
        if factory is None:
            converted[key] = list(parsed)
        else:
            converted[key] = [factory(**item) for item in parsed]

    return converted


def _coerce_record(profile_type: str, record: dict[str, Any]) -> Any:
    normalized = {
        key: _normalize_scalar(value)
        for key, value in record.items()
    }
    normalized = _deserialize_nested(profile_type, normalized)

    profile_class = PROFILE_CLASSES[profile_type]
    float_fields = FLOAT_FIELDS.get(profile_type, set())
    int_fields = INT_FIELDS.get(profile_type, set())
    kwargs: dict[str, Any] = {}
    for field in fields(profile_class):
        value = normalized.get(field.name)
        if value is None:
            continue
        if field.name in int_fields:
            kwargs[field.name] = int(value)
        elif field.name in float_fields:
            kwargs[field.name] = float(value)
        else:
            kwargs[field.name] = value
    return profile_class(**kwargs)


def _write_bundle(bundle: dict[str, list[dict[str, Any]]], root: str | Path = PROFILE_ROOT) -> None:
    for profile_type, key in PROFILE_TYPE_KEYS.items():
        saver = PROFILE_SAVERS[profile_type]
        for record in bundle.get(key, []):
            saver(_coerce_record(profile_type, record), root=root)


def export_profiles_json(output_path: str | Path, root: str | Path = PROFILE_ROOT) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_bundle_from_root(root=root), indent=2), encoding="utf-8")
    return path


def import_profiles_json(input_path: str | Path, root: str | Path = PROFILE_ROOT) -> None:
    bundle = json.loads(Path(input_path).read_text(encoding="utf-8"))
    _write_bundle(bundle, root=root)


def export_profiles_csv(output_dir: str | Path, root: str | Path = PROFILE_ROOT) -> list[Path]:
    output_dir_path = Path(output_dir)
    output_dir_path.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []

    for profile_type, key in PROFILE_TYPE_KEYS.items():
        records = [_serialize_record_for_tabular(record) for record in _bundle_from_root(root=root)[key]]
        fieldnames = [field.name for field in fields(PROFILE_CLASSES[profile_type])]
        path = output_dir_path / f"{key}.csv"
        with path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for record in records:
                writer.writerow({name: record.get(name, "") for name in fieldnames})
        written.append(path)

    return written


def import_profiles_csv(input_dir: str | Path, root: str | Path = PROFILE_ROOT) -> None:
    input_dir_path = Path(input_dir)
    bundle: dict[str, list[dict[str, Any]]] = {key: [] for key in PROFILE_TYPE_KEYS.values()}

    for profile_type, key in PROFILE_TYPE_KEYS.items():
        csv_path = input_dir_path / f"{key}.csv"
        if not csv_path.exists():
            continue
        with csv_path.open("r", newline="", encoding="utf-8") as handle:
            reader = csv.DictReader(handle)
            bundle[key] = [dict(row) for row in reader]

    _write_bundle(bundle, root=root)


def export_profiles_excel(output_path: str | Path, root: str | Path = PROFILE_ROOT) -> Path:
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    first_sheet = True
    bundle = _bundle_from_root(root=root)
    for profile_type, key in PROFILE_TYPE_KEYS.items():
        if first_sheet:
            sheet = workbook.active
            sheet.title = key
            first_sheet = False
        else:
            sheet = workbook.create_sheet(title=key)

        fieldnames = [field.name for field in fields(PROFILE_CLASSES[profile_type])]
        sheet.append(fieldnames)
        for record in bundle[key]:
            serialized = _serialize_record_for_tabular(record)
            sheet.append([serialized.get(name, "") for name in fieldnames])

    workbook.save(path)
    return path


def import_profiles_excel(input_path: str | Path, root: str | Path = PROFILE_ROOT) -> None:
    workbook = load_workbook(filename=Path(input_path))
    bundle: dict[str, list[dict[str, Any]]] = {key: [] for key in PROFILE_TYPE_KEYS.values()}

    for key in PROFILE_TYPE_KEYS.values():
        if key not in workbook.sheetnames:
            continue
        sheet = workbook[key]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(cell) for cell in rows[0]]
        records = []
        for row in rows[1:]:
            if row is None or all(cell in (None, "") for cell in row):
                continue
            records.append(
                {
                    header: row[idx] if idx < len(row) else ""
                    for idx, header in enumerate(headers)
                }
            )
        bundle[key] = records

    _write_bundle(bundle, root=root)
